from openpyxl import load_workbook
import requests
import json


#1、打开一个excel,进入工作薄  wb 工作薄 对象
wb = load_workbook("test.xlsx")

#2、选择工作薄当中一个表单
sh = wb["Sheet"]

#3、读取某一个单元格的值
text = sh.cell(row=2,column=4).value
print("获取到的数据为：",text)

#获取当前文件的总行数
print(sh.max_row)
print(sh.max_column)

# #修改单元格的值
# sh.cell(row=6,column=4).value = "hehehehehehhe"
# #多次修改
#
# #修改的数据 -- 只有保存才能生效 -- 整个工作薄的保存
# wb.save("testdatas.xlsx")

#==========================================================================================
# 将接口请求的过程 定义成函数 -- 只考虑get请求、post请求
def send_reqeust_and_get_resp(method, url, datas):
    if method == "get":
        resp = requests.get(url, datas)
    else:
        resp = requests.post(url, datas)

    return [resp.status_code, resp.text]


#遍历excel文件中的所有行。每获取到一行数据，则发送一次接口请求。
#因为一行数据中有 url、method、req_datas
#row 代表每一行的行号
for index in range(2,sh.max_row +1):
    #获取每一行的测试数据
    method = sh.cell(row=index,column=3).value
    url = sh.cell(row=index,column=4).value
    req_data = sh.cell(row=index,column=5).value
    #期望结果
    expected_data = sh.cell(row=index,column=6).value
    print(type(req_data))
    print("请求方法：",method,"请求地址：",url)
    print("请求的数据：",req_data)
    #请求数据是字符串，需要转换成字典对象  json.loads()
    #发送接口请求
    result = send_reqeust_and_get_resp(method,url,json.loads(req_data))
    print(result)
    print("========================================")
    #将结果写入excel中 == 按行写入
    sh.cell(row=index,column=7).value = result[1]
    #如果期望结果与实际结果相等，我们就写True，否则为False
    if result[1] == expected_data:
        sh.cell(row=index, column=8).value = "True"
    else:
        sh.cell(row=index, column=8).value = "False"

#保存
wb.save("test.xlsx")