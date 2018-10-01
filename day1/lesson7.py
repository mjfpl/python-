__author__ = 'computer'
from openpyxl import load_workbook
import requests
from openpyxl import Workbook
import json
def request_prox(method,url,datas):
    if method == "get":
        resp =requests.get(url,datas)
    else:
        resp = requests.post(url,datas)
    return [resp.status_code,resp.text]
# wb = Workbook("test1.xlsx") 创建一个excel表格
# ws = wb.active
# wb.save("test1.xlsx")
wb = load_workbook("test1.xlsx")
ws = wb["Sheet"]
#给excel进行赋值
#reult = ws.cell(row=1,column=1).value = 20
# reult2 = ws.cell(row=2,column=5).value
# print(reult2)
# # 获取excel的行跟列
# row_num = ws.max_row
# print(row_num)
# column_num = ws.max_column
# print(column_num)
# wb.save("test1.xlsx")

for item in range(2,ws.max_row+1):
     method = ws.cell(row=item,column=3).value
     url = ws.cell(row=item,column=4).value
     datas = ws.cell(row=item,column=5).value
     print(json.loads(datas))
     result = request_prox(method,url,json.loads(datas))
     print(result)