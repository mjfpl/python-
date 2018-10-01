__author__ = 'computer'
from openpyxl import load_workbook
import requests
from json import loads
# 首先读取工作簿
wb = load_workbook("test1.xlsx")
# 读取sheet
sheet = wb["Sheet"]
# row_num = sheet.max_row
# column_num = sheet.max_column
# print(row_num,column_num)
# 定义发送一个请求
#sheet.cell(row=item,column=6).value = result[1]
def request_fasong(method,url,datas):
    if method == "get":
        resp =requests.get(url,datas)
    else:
         resp =requests.get(url,datas)
    return [resp.status_code,resp.text]

for item in range(2,sheet.max_row+1):
    method = sheet.cell(row=item, column=3).value
    url = sheet.cell(row=item, column=4).value
    datas = sheet.cell(row=item, column=5).value
    result = request_fasong(method, url, loads(datas))
    print(result)
    sheet.cell(row=item,column=6).value = result[1]
#对excel进行保存的时候一定要把excel表关闭
wb.save("test1.xlsx")