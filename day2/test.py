
# from openpyxl import Workbook
# wb = Workbook()
# ws = wb.active
# ws1 = wb.create_sheet()
# ws2 = wb.create_sheet(0)
# ws.title = "New Title"

from openpyxl import Workbook
wb = Workbook()
# 激活 worksheet
ws = wb.active
# 数据可以直接分配到单元格中
ws['A1'] = 42
ws['B1'] = 2
# 可以附加行，从第一列开始附加
ws.append([1, 2, 3])
# Python 类型会被自动转换
import datetime
ws['A3'] = datetime.datetime.now().strftime("%Y-%m-%d")
ws1 = wb.create_sheet("test") # 新建一个工作表
ws2 = wb.create_sheet("test1",0)

# 保存文件
#wb.save("sample.xlsx")
print(wb.sheetnames)  # 获取每个工作表的名字
for item in wb:
    print(item.title) # 遍历工作簿的名字

# 操作数据  访问单元格
c = ws["A2"]  # 这样只能够获取单元格所在的工作表的坐标
d = ws.cell(row=1, column=2, value=20)
for i in range(1,101):
    for j in range(1,101):
        ws.cell(row=i, column=j)
ws["B1"] = 20