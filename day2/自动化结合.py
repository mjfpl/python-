__author__ = 'computer'
import requests
from openpyxl import load_workbook
import json
def request_res(method, url, datas):
    if method == "get":
        resp = requests.get(url,datas)
    else:
        resp =requests.get(url,datas)
    return [resp.status_code,resp.text]
wb = load_workbook("test1.xlsx")
ws =wb["Sheet"]
for item in range(2,ws.max_row+1):
    method = ws.cell(row=item, column=3).value
    url = ws.cell(row=item, column=4).value
    datas = ws.cell(row=item, column=5).value
    response = request_res(method,url,json.loads(datas))
    print(response)