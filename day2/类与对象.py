__author__ = 'computer'
import requests
from openpyxl import load_workbook
from json import loads
wb = load_workbook("test.xlsx")
sh = wb["Sheet"]
# requests 请求模块
class request1:
    #初始化实体类
    def __init__(self,method, url, data):
        self.method= method
        self.url= url
        self.data= data
    #定义函数
    def  request_qingqiu(self, method, url, data):
        if method == "get":
            resp = requests.get(url,data)
        else:
            resp = requests.post(url,data)
        return resp.text

for item in range(2,sh.max_row+1):
    #对类进行赋值
    #获取请求的方法
    request1.method = sh.cell(row=2,column=3).value
    print(request1.method)
    #获取请求的地址
    request1.url = sh.cell(row=2,column=4).value
    print(request1.url)
    #获取请求的参数
    request1.data = sh.cell(row=2,column=5).value
    data = loads(request1.data)
    print(type(data))
    # print(request.data)
    # print(type(request.data))
    # print(data)
    #调用函数 并传参
    # r = requests.get(request.url,data)
    # print(r.text)
    request1.request_qingqiu(request1.method, request1.url, data)
wb.save("test.xlsx")